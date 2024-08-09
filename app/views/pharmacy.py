from app.views import *
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from app.models import Pharmacy

@method_decorator(csrf_exempt, name="dispatch")
class PharmacyCreateView(CreateView):
    model = Pharmacy
    form_class = PharmacyForm
    template_name = 'create.html'
    success_url = reverse_lazy('successfully_created')

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['title'] = 'Создание аптек'
        return context

def export_pharmacies_to_excel(request):
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Pharmacies"

    # Define the headers
    headers = [
        "Юридическое название", 
        "ФИО фармацевта 1", 
        "ФИО фармацевта 2", 
        "Ответственный", 
        "Контакты", 
        "Контакт ответственного", 
        "Район"
    ]

    # Write the headers to the first row
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet[f"{column_letter}1"] = header

    # Write data rows
    for row_num, pharmacy in enumerate(Pharmacy.objects.all(), 2):
        worksheet[f"A{row_num}"] = pharmacy.title
        worksheet[f"B{row_num}"] = pharmacy.name
        worksheet[f"C{row_num}"] = pharmacy.name2
        worksheet[f"D{row_num}"] = pharmacy.responsible
        worksheet[f"E{row_num}"] = pharmacy.contact
        worksheet[f"F{row_num}"] = pharmacy.responsible_contact
        worksheet[f"G{row_num}"] = pharmacy.region.title if pharmacy.region else ""  # Assuming region has a `name` attribute

    # Set the width of the columns for better readability
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="pharmacies.xlsx"'
    workbook.save(response)
    return response
