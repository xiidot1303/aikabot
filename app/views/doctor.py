from app.views import *
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from app.views import *
from openpyxl import Workbook
from app.services.doctor_service import Doctor
from openpyxl.utils import get_column_letter

@method_decorator(csrf_exempt, name="dispatch")
class DoctorCreateView(CreateView):
    model = Doctor
    form_class = DoctorForm
    template_name = 'create.html'
    success_url = reverse_lazy('successfully_created')

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context['title'] = 'Создание доктора'
        context['form'].fields['fillial'].initial = self.request.GET.get('fillial')
        context['form'].fields['fillial'].label = ""
        return context


def export_doctors_view(request):
    # Create a new workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Doctors"

    # Define the headers
    headers = ["ФИО врача", "Контакты", "Направление", "Место работы", "Район"]

    # Write the headers to the first row
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        worksheet[f"{column_letter}1"] = header

    # Write data rows
    for row_num, doctor in enumerate(Doctor.objects.all(), 2):
        worksheet[f"A{row_num}"] = doctor.name
        worksheet[f"B{row_num}"] = doctor.contact
        worksheet[f"C{row_num}"] = doctor.direction
        worksheet[f"D{row_num}"] = doctor.workplace
        worksheet[f"E{row_num}"] = doctor.region.title if doctor.region else ""  # Assuming region has a `name` attribute

    # Set the width of the columns for better readability
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="doctors.xlsx"'
    workbook.save(response)
    return response