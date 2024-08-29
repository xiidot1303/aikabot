from app.views import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import messages
from app.services.visit_service import filter_visits_by_date_range

async def export_visits_view(request):
    start_date = request.GET.get('datetime__range__gte')
    end_date = request.GET.get('datetime__range__lte')
    start_date = "10.10.2020" if not start_date else start_date
    end_date = "10.10.2100" if not end_date else end_date

    # Создаем новый Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Посещения"

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 17
    ws.column_dimensions['Q'].width = 60
    # Устанавливаем стандартную ширину для колонок от C до M
    standard_width = 15  # Стандартная ширина
    for col in range(ord('C'), ord('Q')):  # C (67) to M (77)
        ws.column_dimensions[chr(col)].width = standard_width

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws["A1"] = "Мед представитель"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=6)
    ws["C1"] = "Врач"
    ws["C1"].font = Font(size=14, bold=True)
    ws["C1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=12)
    ws["G1"] = "Аптека"
    ws["G1"].font = Font(size=14, bold=True)
    ws["G1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=1, start_column=13, end_row=1, end_column=13)
    ws["M1"] = "Партнёр"
    ws["M1"].font = Font(size=14, bold=True)
    ws["M1"].alignment = Alignment(horizontal="center")


    # Записываем заголовки
    headers = [
        "Имя", "Номер телефона", 
        # doctor
        "ФИО врача", "Контакты", "Направление", "Место работы", 
        # pharmacy
        "Ответственный", "Контакт ответственного", "ФИО фармацевта 1",
        "ФИО фармацевта 2", "Контакты", "Юридическое название",
        # partner
        "Имя",

        "Комментария", "Дата", "Тип визита", "Расположение"
        ]
    ws.append(headers)

    # Записываем данные
    async for visit in filter_visits_by_date_range(start_date, end_date):
        doctor: Doctor = await visit.get_doctor()
        pharmacy: Pharmacy = await visit.get_pharmacy()
        partner: Partner = await visit.get_partner()
        ws.append([
            (await visit.get_bot_user()).name, (await visit.get_bot_user()).phone, 
            # doctor
            doctor.name if doctor else "", doctor.contact if doctor else "",
            doctor.direction if doctor else "", doctor.workplace if doctor else "", 
            # pharmacy
            pharmacy.responsible if pharmacy else "", pharmacy.responsible_contact if pharmacy else "",
            pharmacy.name if pharmacy else "", pharmacy.name2 if pharmacy else "",
            pharmacy.contact if pharmacy else "", pharmacy.title if pharmacy else "",
            # partner
            partner.name if partner else "",

            visit.comment, 
            visit.datetime.strftime("%d.%m.%Y %H:%M"), visit.get_type_display(), visit.location
            ])

    # Define border style
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to the entire table
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_style

    # color the fields
    for row in range(1, ws.max_row + 1):
        for col in range(1, 3):  # Columns A, B
            ws.cell(row=row, column=col).fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        for col in range(3, 7):  # Columns C, D, E, F, G
            ws.cell(row=row, column=col).fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        for col in range(7, 13):  # Columns G, H, I, J, K, L 
            ws.cell(row=row, column=col).fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        for col in range(13, 14):  # Column M
            ws.cell(row=row, column=col).fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
        for col in range(14, 18):  # Columns N, O, P, Q
            ws.cell(row=row, column=col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Подготавливаем HTTP ответ с Excel файлом
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=visits.xlsx'
    wb.save(response)
    return response